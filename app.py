import streamlit as st
import torch
import torch.nn as nn
import pandas as pd
from torch.nn.utils.rnn import pack_padded_sequence
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =====================
# Load Model
# =====================
class LSTM(nn.Module):
    def __init__(self, vocab_size, embed_size, hidden_dim):
        super().__init__()
        self.embedding = nn.Embedding(vocab_size, embed_size, padding_idx=0)
        self.lstm = nn.LSTM(embed_size, hidden_dim, batch_first=True)
        self.fc = nn.Linear(hidden_dim, 1)

    def forward(self, x, lengths):
        x = self.embedding(x)
        packed = pack_padded_sequence(x, lengths, batch_first=True, enforce_sorted=False)
        _, (hidden, _) = self.lstm(packed)
        return self.fc(hidden[-1])

# Load saved model
checkpoint = torch.load("junk_model.pth", map_location="cpu")
vocab = checkpoint["vocab"]

model = LSTM(len(vocab), 128, 128)
model.load_state_dict(checkpoint["model_state_dict"])
model.eval()

# =====================
# Encode Function
# =====================
def encode(text):
    text = str(text).lower().strip()
    words = text.split()
    return [vocab.get(word, vocab["<UNK>"]) for word in words]

def predict_text(text):
    seq = encode(text)
    if len(seq) == 0:
        seq = [vocab["<UNK>"]]

    seq_tensor = torch.tensor(seq).unsqueeze(0)
    length = [len(seq)]

    with torch.no_grad():
        output = model(seq_tensor, length)
        prob = torch.sigmoid(output)
        pred = (prob > 0.5).int().item()

    return pred, prob.item()

# =====================
# Streamlit UI
# =====================

st.set_page_config(
    page_title="Junk Detection Tool",
    page_icon="🧹",
    layout="wide"
)

# Custom CSS Styling
st.markdown("""
    <style>
    .main {
        background-color: #f5f7fa;
    }

    .title {
        font-size: 36px;
        font-weight: bold;
        color: #2c3e50;
        text-align: center;
        padding: 10px;
    }

    .subtitle {
        font-size: 18px;
        text-align: center;
        color: #555;
        margin-bottom: 30px;
    }

    .stButton>button {
        background-color: #ff4b4b;
        color: white;
        font-size: 16px;
        padding: 10px 24px;
        border-radius: 8px;
        border: none;
    }

    .stButton>button:hover {
        background-color: #e03e3e;
        color: white;
    }

    .success-box {
        background-color: #d4edda;
        padding: 15px;
        border-radius: 8px;
        color: #155724;
        font-weight: 500;
    }
    </style>
""", unsafe_allow_html=True)


st.title("📊 Junk Detection in Excel")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file is not None:
    df = pd.read_excel(uploaded_file)

    column = st.selectbox("Select column to analyze", df.columns)

if st.button("Analyze"):

    wb = load_workbook(uploaded_file)
    ws = wb.active

    red_fill = PatternFill(start_color="FF9999",
                           end_color="FF9999",
                           fill_type="solid")

    # Find column index
    col_index = list(df.columns).index(column) + 1  # Excel starts from 1

    for idx, value in enumerate(df[column]):
        pred, prob = predict_text(value)
        print(value, "->", pred, prob)

        if prob >= 0.9:
            ws.cell(row=idx+2, column=col_index).fill = red_fill


    output_file = "highlighted_output.xlsx"
    wb.save(output_file)

    with open(output_file, "rb") as f:
        st.download_button(
            "Download Result File",
            f,
            file_name="junk_marked.xlsx"
        )

    st.success("Analysis complete ✅")


